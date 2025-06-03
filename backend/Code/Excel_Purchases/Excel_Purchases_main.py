from openpyxl import load_workbook
from pathlib import Path
from openpyxl import load_workbook


def find_first_empty_row_purchuses(ws):
    for row in range(4, ws.max_row + 2):
        if ws.cell(row, 2).value is None:
            return row
    return ws.max_row + 1

def copy_CreoMate_to_Purchases(Excel_path, Purchases_Excel_path, ws_BOM, ws_Purchases):
    start_row = find_first_empty_row_purchuses(ws_Purchases)

    print("Start copying from row:", start_row, flush=True)

   

    for row in range(2, ws_BOM.max_row + 1):
        for col in range(2, 10):
            value = ws_BOM.cell(row, col).value
            ws_Purchases.cell(start_row, col).value = value
            
        start_row += 1
    

    


def main(Excel_path, Purchases_Excel_path):
    wb_BOM = load_workbook(Excel_path)
    wb_Purchases = load_workbook(Purchases_Excel_path)
    ws_BOM = wb_BOM.active
    ws_Purchases = wb_Purchases.active


    copy_CreoMate_to_Purchases(Excel_path, Purchases_Excel_path, ws_BOM, ws_Purchases)

    wb_Purchases.save(Purchases_Excel_path)
    
