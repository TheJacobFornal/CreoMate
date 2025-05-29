from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill

curr_dir = Path(__file__).parent
Check_dir = curr_dir.parent
code_dir = Check_dir.parent
app_dir = code_dir.parent
BOM_dir = app_dir / "BOM"
temp_dir = BOM_dir / "template"
bom_path = BOM_dir / "bom_ready.txt"
excel_path = BOM_dir / "bom_ready.xlsx"

def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill

def check_profile(ws, row):                                                 # compare number H and length of profile kol. A wit kol B - grey
    Number_Index = 2
    Creo_Index = 1

    number_val = ws.cell(row, Number_Index).value
    creo_val = ws.cell(row, Creo_Index).value

    number_creo = creo_val.split("-")[0]

    if number_creo != number_val:
        color_row(ws, row, True, "A1A1A1")
    else:
        color_row(ws, row, False)

def modify_left_duplicate(ws, row_left, creo_left_number):
    Quantity_Index = 4
    Creo_Index = 1
    Uwagi_Index = 9
    left_quantity = ws.cell(row_left, Quantity_Index).value

    right_name = creo_left_number[:-1]

    right = False
    for row in range(1, ws.max_row + 1):                    # find right component
        creo_right_value = ws.cell(row, Creo_Index).value
        if creo_right_value and creo_right_value.__contains__(right_name) and not creo_right_value.__contains__(creo_left_number):           # there is right component
            right_quantity = ws.cell(row, Quantity_Index).value
            if not str(right_quantity).__contains__("+"):
                quantity = str(right_quantity).strip() + " + " + str(left_quantity).strip() + "L"
                ws.cell(row, Quantity_Index).value = quantity                   # right component
                color_row(ws, row_left, True, "FF0095")
            right = True
            break

    if not right:                                                                                                       # there isn't right component
        left_quantity = ws.cell(row_left, Quantity_Index).value

        if not str(left_quantity).__contains__("+"): # no duplicates
            quantity = "0+ " + str(left_quantity).strip() + "L"
            ws.cell(row_left, Quantity_Index).value = quantity

            Note = "Wykonać w lustrze!"
            ws.cell(row_left, Uwagi_Index).value = Note

            color_row(ws, row_left, True, "42FF48")


def main():
    Name_Index = 3
    Type_Index = 5
    Creo_Index = 1
    wb = load_workbook(excel_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        Name_value = ws.cell(row, Name_Index).value
        Type_value = ws.cell(row, Type_Index).value
        creo_val = ws.cell(row, Creo_Index).value

        if Name_value is not None:
            Upper_name = Name_value.upper()
            ws.cell(row, Name_Index).value = Upper_name                                            # upper letter in kol. Name
            if Name_value.__contains__("PROFIL_") and Type_value.__contains__("H"):                # Profile with dimensions Type: H
                check_profile(ws, row)
            else:
                Numer_val = ws.cell(row, 2).value                                                  # Number contain "_" - brown
                if Numer_val is not None and  Numer_val.__contains__("_"):
                    color_row(ws, row, True, "D3A6FF")

                number_creo = creo_val.split("-")[0]                                               # Left element
                if number_creo.__contains__("L") and Type_value != "H":
                    modify_left_duplicate(ws, row, number_creo)




    wb.save(excel_path)