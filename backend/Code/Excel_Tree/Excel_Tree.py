from pathlib import Path
from openpyxl import load_workbook, Workbook
import os
from openpyxl.styles import Font
from pathlib import Path
from openpyxl.styles import PatternFill


## Small utilities ##
def get_type(creoName: str) -> str:
    return str(creoName).split("-")[-1]


def contain_digit(word) -> bool:
    return any(ch.isdigit() for ch in str(word))


def color_row(
    ws, row_num: int, highlight: bool, color: str = "FFFF00", wrong_counter: int = 0
) -> int:
    """Color columns A..F of a row. If highlight=True, increment wrong_counter."""
    if highlight:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        wrong_counter += 1
    else:
        fill = PatternFill(fill_type=None)

    for col in range(1, 7):
        ws.cell(row=row_num, column=col).fill = fill

    return wrong_counter


def remove_cell_color(ws):
    """Clear fill from all cells."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type is not None:
                cell.fill = PatternFill(fill_type=None)


def clear_producent(ws, row: int):
    """Normalize producer name in column F (6)."""
    name = ws.cell(row, 6).value
    if not name:
        return
    s = str(name)
    if "kat." in s.lower() or "kat" in s.lower():  # remove "kat." or "kat"
        s = s.lower().replace("kat.", "").replace("kat", "")
    s = s.lower().strip().capitalize()
    ws.cell(row, 6).value = s


## Check specific Type Product ##
def check_handlowe_normalia(ws, row: int, wrong_counter: int) -> int:
    """
    If 'producent' (col F) is missing:
      - if col A contains 'Assembly' -> delete row
      - otherwise -> highlight red.
    Else -> clear producent.
    """
    producent_val = ws.cell(row, 6).value
    first_col_val = ws.cell(row, 1).value or ""

    if not producent_val:  # missing producent
        if "Assembly" in str(first_col_val):
            ws.delete_rows(row)
            return wrong_counter
        else:
            wrong_counter = color_row(ws, row, True, "FF0000", wrong_counter)
    else:
        clear_producent(ws, row)

    return wrong_counter


def handlowe_number(ws, row: int, wrong_counter: int) -> int:
    """Highlight Hx/Nx items (no removal)."""
    return color_row(ws, row, True, "FFFF00", wrong_counter)


def remove_h_number(ws, row: int):
    """Delete the row (used when removing Hx/Nx items)."""
    ws.delete_rows(row)


## Main function ##
def main(Excel_path: str, remove_h_items: bool):
    wb = load_workbook(Excel_path)
    ws = wb.active

    # start fresh
    remove_cell_color(ws)
    wrong_counter = 0

    # iterate bottom-up so deletions don't skip rows
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row, 2).value  # column B
        val_str = "" if val is None else str(val)

        # if B cell doesn't look like "part-xxx"
        if len(val_str.split("-")) < 2:
            wrong_counter = color_row(ws, row, True, "00FFB7", wrong_counter)
            continue

        type_str = get_type(val_str)
        tl = type_str.lower()

        # Handlowy (H) or Normalia (N)
        if ("h" in tl) or ("n" in tl):
            if contain_digit(tl):  # e.g., H1, H2, N3
                if remove_h_items:
                    remove_h_number(ws, row)
                else:
                    wrong_counter = handlowe_number(ws, row, wrong_counter)
            else:
                wrong_counter = check_handlowe_normalia(ws, row, wrong_counter)

    # Save before opening the file
    wb.save(Excel_path)

    # keeping your original return signature
    return ws.max_row + 1, wrong_counter


def copy_data_to_template(source_path, dest_path, start_row=4):
    """
    Copy values from source (active sheet) B2:I{max} into dest (active sheet),
    starting at the first empty row >= start_row (scanning columns B..I).
    Saves back to the same destination template (xlsx, no macros).
    """
    # Load workbooks
    src_wb = load_workbook(source_path, data_only=True)
    dst_wb = load_workbook(dest_path)  # your new template (.xlsx)
    src_ws = src_wb.active
    dst_ws = dst_wb.active

    #

    # Copy values from source rows 2..max, cols B..I, into dest starting at write_row

    for r in range(1, src_ws.max_row + 1):

        for c in range(1, 7):  # B..I
            dst_ws.cell(r + 3, c).value = src_ws.cell(r, c).value

    # Save back to the template file
    dst_wb.save(dest_path)

    ### Phase 3 ###


from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font


def link_path(ws, row, Tree_PDF_dir, Tree_DWG_dir) -> bool:
    """Insert a hyperlink to <base_name>.pdf in column G; return True if missing."""
    base_name = ws.cell(row, 2).value  # column B
    if not base_name:
        return True  # nothing to link

    pdf_name = f"{base_name}.pdf"
    file_path_pdf = Tree_PDF_dir / pdf_name
    dwg_name = f"{base_name}.dwg"
    file_path_dwg = Tree_DWG_dir / dwg_name

    missing = False

    print(file_path_pdf, file_path_dwg)

    if not file_path_dwg.exists():
        color_row(ws, row, True, "FFA500")  # if you want to mark missing
        missing = True

    if file_path_pdf.exists():
        cell = ws.cell(row=row, column=7)  # column G
        cell.value = pdf_name  # text shown
        cell.hyperlink = str(file_path_pdf)  # clickable path
        cell.font = Font(color="0000FF", underline="single")
        return missing
    else:
        color_row(ws, row, True, "00B0F0")  # if you want to mark missing
        return True


def link_drowings(excel_path, Tree_DWG_dir, Tree_PDF_dir):
    wb = load_workbook(excel_path)
    ws = wb.active

    VALID_TYPES = {"H", "P", "L", "F", "T", "W", "O", "S", "N", "D"}
    counter_missing = 0
    counter_drawings = 0

    for row in range(4, ws.max_row + 1):
        creo_name = ws.cell(row, 2).value
        if not creo_name:
            continue
        typ = str(creo_name).split("-")[-1].upper()  # normalize

        if typ in VALID_TYPES:
            counter_drawings += 1  # ✅ increment
            if link_path(ws, row, Tree_PDF_dir, Tree_DWG_dir):
                counter_missing += 1  # ✅ increment

    wb.save(excel_path)
    print(f"Linked drawings. Total: {counter_drawings}, missing: {counter_missing}")

    correct = counter_drawings - counter_missing

    percentage = correct / counter_drawings * 100

    if counter_drawings != 0:
        percentage = round((correct / counter_drawings) * 100)
        text = f"{correct}/{counter_drawings} ({percentage}%) - znaleziono"
    else:
        text = "0/0 (0%)"

    return text


def remove_colors(excel_path: Path):
    if Path(excel_path).is_file():
        wb = load_workbook(excel_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            for cell in row:
                if cell.fill and cell.fill.fill_type is not None:
                    cell.fill = PatternFill(fill_type=None)

        wb.save(excel_path)
        print(f"All fills cleared from row 4 down in {excel_path}")
    else:
        print("File not found:", excel_path)


if __name__ == "__main__":
    remove_colors(Path(r"D:\Creo_Ustawienia\Programiki\CreoMate\BOM CreoMate.xlsx"))
    link_drowings(
        Path(
            r"D:\Creo_Ustawienia\Programiki\CreoMate\Dokumancja\Gotowa Dokumentacja.xlsx"
        ),
        Path(r"D:\Creo_Ustawienia\Programiki\CreoMate\Dokumancja\DWG"),
        Path(r"D:\Creo_Ustawienia\Programiki\CreoMate\Dokumancja\PDF"),
    )
    os.startfile(Path(r"D:\Creo_Ustawienia\Programiki\CreoMate\BOM CreoMate.xlsx"))
