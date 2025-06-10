from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wrong_counter = 0
Zakupy = False
drowings_dir = None

def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill

def check_if_spawane_00(name):
    name = name.split("-")[1]
    name = str(name)
    if name[-2:] == "00":
        print("spawane: ", name)
        return True
    else:
        return False

def check_dir(file_name, row, ws):
    global drowings_dir
    global wrong_counter    
    path = drowings_dir / file_name
    if not path.is_dir():
        color_row(ws, row, True, "F8DF00")      # yellow
        wrong_counter += 1
    else:
        color_row(ws, row, False)




def divide_elem(ws):
    global wrong_counter
    global drowings_dir
    global Zakupy
    
    Typ_index = 5
    Creo_index = 1
    for row in range(1, ws.max_row + 1):
        Typ_value = ws.cell(row, Typ_index).value
        if Typ_value == "F" or Typ_value == "L" or Typ_value == "W" or Typ_value == "T" or Typ_value == "O" or Typ_value == "D":
            if Zakupy:
                file_name = file_name_from_zakupy(row, ws)
            else:
                file_name = ws.cell(row, Creo_index).value
             
            if check_if_spawane_00(file_name):
                check_dir(file_name, row, ws)
            else:
                check_elem(file_name, row, ws)
                
                
                
def file_name_from_zakupy(row, ws):
    global Zakupy
    
    number_val = ws.cell(row, 2).value
    name_val = ws.cell(row, 3).value
    type_val = ws.cell(row, 5).value
    

    if number_val is not None and name_val is not None:
        number_val = str(number_val).strip()
        parts = number_val.split(".")
        project_name = parts[0] + "_" + parts[1] + parts[2] + parts[3] + parts[4]
        
        file_name = project_name + "-" + name_val.lower().strip() + "-" + type_val.lower()
    
    return file_name
        


def check_elem(file_name, row, ws):                                 #check if file exist in dir (elem)
    global wrong_counter
    global Zakupy
    global drowings_dir
        
        
    pdf_name = file_name + ".pdf"                               #pdf file
    file_path_pdf = drowings_dir / pdf_name

    bool = False

    if not file_path_pdf.exists():
        color_row(ws, row, True, "00B0F0")         
        bool = True
        wrong_counter += 1
        print("PDF not found: ", file_path_pdf, flush=True)



    stp_name = file_name + ".stp"                               #stp file
    file_path_stp = drowings_dir / stp_name
    if not file_path_stp.exists():
        color_row(ws, row, True, "00B0F0")         
        if not  bool:
            wrong_counter += 1
            print("stp not found: ", stp_name, flush=True)
            bool = True


    dwg_name = file_name + ".dwg"                               #dwg file
    file_path_dwg = drowings_dir / dwg_name
    if not file_path_dwg.exists():
        color_row(ws, row, True,  "00B0F0")         
        if not bool:
            print("dwgt not found: ", dwg_name, flush=True)
            wrong_counter += 1


def main(Excel_path, folder, Zakupy_bool=False):
    global Zakupy
    global wrong_counter
    global drowings_dir
    
    Zakupy = Zakupy_bool
    wrong_counter = 0
    wb = load_workbook(Excel_path)
    ws = wb.active
    drowings_dir = Path(folder)
    divide_elem(ws)

    wb.save(Excel_path)
    return wrong_counter