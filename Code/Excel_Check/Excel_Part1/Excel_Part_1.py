from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill

curr_dir = Path(__file__).parent
Check_dir = curr_dir.parent
code_dir = Check_dir.parent
app_dir = code_dir.parent
BOM_dir = app_dir / "BOM"
temp_dir = BOM_dir / "template"
bom_path = BOM_dir / "bom_ready.txt"
excel_path = BOM_dir / "bom_ready.xlsx"

def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill

def contain_num(text):
    return any(char.isdigit() for char in text)

def remove_dash_Type(ws):                                                           # remove - and remove H1, H2, Hx
    TypeIndex = 5
    for row in range(1, ws.max_row + 1):
        Type_value = ws.cell(row, TypeIndex).value
        if Type_value is not None and not Type_value.isdigit():
            color_row(ws, row, False) # has Type
            removed_dash = Type_value.replace("-", "")
            ws.cell(row, TypeIndex).value = removed_dash

            if removed_dash.__contains__("H") and contain_num(removed_dash):        # H2, h3, h...  - yellow
                color_row(ws, row, True)
        else:
            color_row(ws, row, True,"8CBCF4")                              # Empty type - light blue


def check_handlowe(ws, row):
    Producent_index = 9                      # Producent
    Nazwa_index = 3
    Numer_index = 2

    Producent_cell = ws.cell(row, Producent_index)
    Nazwa_val = ws.cell(row, Nazwa_index).value
    Numer_val = ws.cell(row, Numer_index).value

    def_numer = "nr. katalogowy"
    def_nazwa = "nazwa katalogowa"
    def_producent = "Producent"

    if Nazwa_val.strip() == def_nazwa or Numer_val.strip() == def_numer or Producent_cell.value == def_producent:       #check czy default nazwy - nie uzupełnione relacje kol. B, C, I
        color_row(ws, row, True, "F76700")
        return

    if not Producent_cell.value:                                                                     # nie ma producenta  - red
        color_row(ws, row, True, "FF0000")
    else:
        color_row(ws, row, False)                                                               # jest producent - empty

def check_production(ws, row):
    Material_index = 6
    Cieplna_index = 7
    Powierzchnia_index = 8

    Material_cell = ws.cell(row, Material_index)
    Cieplna_cell = ws.cell(row, Cieplna_index)
    Powierzchnia_cell = ws.cell(row, Powierzchnia_index)

    if not Material_cell.value or not Cieplna_cell.value or not Powierzchnia_cell.value:               # Produkiowane nie materiału i obróbek
        color_row(ws, row, True, "ABA200")
    else:
        print(Cieplna_cell.value)
        if Cieplna_cell.value.lower() == "brak":   # zmienia "Brak" na "Brak / None"
            ws.cell(row, Cieplna_index).value  = "Brak / None"
        elif Powierzchnia_cell.value.lower() == "brak":
            ws.cell(row, Powierzchnia_index).value = "Brak / None"
        color_row(ws, row, False)



def main():
    TypeIndex = 5
    wb = load_workbook(excel_path)
    ws = wb.active
    remove_dash_Type(ws)

    for row in range(1, ws.max_row + 1):
        Type_value = ws.cell(row, TypeIndex).value

        if Type_value is not None:
            if Type_value == "H" or Type_value == "N":
                check_handlowe(ws, row)
            elif not Type_value.__contains__("H") and not Type_value.isdigit():
                check_production(ws, row)

    wb.save(excel_path)