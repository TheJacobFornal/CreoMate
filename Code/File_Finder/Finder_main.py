from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


curr_dir = Path(__file__).parent
Code_dir = curr_dir.parent
app_dir = Code_dir.parent
BOM_dir = app_dir / "BOM"
temp_dir = BOM_dir / "template"
excel_path = BOM_dir / "bom_ready.xlsx"

folder_path = Path(r"C:\Users\JakubFornal\Desktop\243201-IX30-4-Wewnetrzne(Skrecanie)")

def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill

def check_if_spawane_00(name):
    name = name.split("_")[1]
    name = str(name.split("-")[0])

    if name[-2:] == "00":
        print("spawane: ", name)
        return True
    else:
        return False

def check_dir(creo_name, row, ws):
    path = folder_path / creo_name
    if not path.is_dir():
        color_row(ws, row, True, "F8DF00")      # yellow
    else:
        color_row(ws, row, False)




def get_creo_name(ws):
    Typ_index = 5
    Creo_index = 1
    for row in range(1, ws.max_row + 1):
        Typ_value = ws.cell(row, Typ_index).value
        if Typ_value == "F" or Typ_value == "L" or Typ_value == "W" or Typ_value == "T" or Typ_value == "O" or Typ_value == "D":
            creo_name = ws.cell(row, Creo_index).value
            if check_if_spawane_00(creo_name):
                check_dir(creo_name, row, ws)
            else:
                check_elem(creo_name, row, ws)


def check_elem(creo_name, row, ws):                                 #check if file exist in dir (elem)
    pdf_name = creo_name + ".pdf"                               #pdf file
    file_path_pdf = folder_path / pdf_name
    if not file_path_pdf.exists():
        color_row(ws, row, True, "FF2A00")          #red
    else:
        color_row(ws, row, False)


    stp_name = creo_name + ".stp"                               #stp file
    file_path_stp = folder_path / stp_name
    if not file_path_stp.exists():
        color_row(ws, row, True, "115BFA")          # blue
    else:
        color_row(ws, row, False)

    dwg_name = creo_name + ".dwg"                               #dwg file
    file_path_dwg = folder_path / dwg_name
    if not file_path_dwg.exists():
        color_row(ws, row, True,  "5A5B5D")         # grey
    else:
        color_row(ws, row, False)

def main():
    wb = load_workbook(excel_path)
    ws = wb.active
    get_creo_name(ws)

    wb.save(excel_path)



