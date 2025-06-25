from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill


# List of hex RGB colors (uppercase, without '#') to match
TAG_COLORS = {
    "00FFB7",  # Mint Aqua
    "FFFF00",  # Bright Yellow
    "FF0000",  # Red
    "F76700",  # Orange
    "ABA200",  # Olive
    "D3A6FF",  # Lavender Violet
    "00B0F0",  # Cyan Blue
    "A1887F",  # Light Brown / Taupe
    "B0BEC5",  # Cool Gray
    "FF3399",  # Hot Pink
    "42FF48",  # Neon Green
    "DDD8B8",  # sandy,
    "379392",  # Dark syjan
}

def remove_tag_colors(ws):
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill and fill.fill_type == "solid":
                color_obj = fill.start_color

                # Get RGB color safely
                if color_obj.type == "rgb" and color_obj.rgb:
                    hex_rgb = color_obj.rgb[-6:].upper()
                    if hex_rgb in TAG_COLORS:
                        cell.fill = PatternFill(fill_type=None)



def main(Excel_path):
    if Excel_path.is_file():
        wb = load_workbook(Excel_path)
        ws = wb.active
        remove_tag_colors(ws)

        wb.save(Excel_path)

def number_of_rows(Excel_path, Zakupy=False):
    if Excel_path.is_file():
        wb = load_workbook(Excel_path)
        ws = wb.active
        counter = 0
        if Zakupy:
            number_index = 2
            for row in range(5, 600):
                value = ws.cell(row, number_index).value
                if value not in ("", None):
                    counter += 1
            return counter
        else:
            return ws.max_row
    return 0


def number_of_rows_drawings(Excel_path):
    if Excel_path.is_file():
        wb = load_workbook(Excel_path)
        ws = wb.active
        counter = 0
        typeINdex = 5
        for row in range(1, ws.max_row + 1):
            type = ws.cell(row, typeINdex).value
            
            if type is not None and (type == "F" or type == "P" or type == "S" or type == "L" or type == "W" or type == "T" or type == "O" or type == "D"):
               
                counter += 1
        return counter
    return counter


def get_max_min_row(ws, zakupy=False):
    global max_row
    global min_row
    max_row = 0
    min_row = 0
    for row in ws.iter_rows(min_row=5, max_col=2, values_only=True):
        if row[1] is not None:
            max_row += 1
    if zakupy:
        max_row += 4
        min_row = 5
    else:
        min_row = 2
        max_row = ws.max_row
    return max_row, min_row







