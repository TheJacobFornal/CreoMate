from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill

curr_dir = Path(__file__).parent
Check_dir = curr_dir.parent
code_dir = Check_dir.parent
app_dir = code_dir.parent
BOM_dir = app_dir / "BOM"
temp_dir = BOM_dir / "template"
bom_path = BOM_dir / "bom_ready.txt"
excel_path = BOM_dir / "bom_ready.xlsx"

def remove_bg_color(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(fill_type=None)

def main():
    if excel_path.is_file():
        wb = load_workbook(excel_path)
        ws = wb.active
        remove_bg_color(ws)

        print("bg color removed")

        wb.save(excel_path)





