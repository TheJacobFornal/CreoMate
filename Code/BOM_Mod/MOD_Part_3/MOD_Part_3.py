from pathlib import Path
from openpyxl import load_workbook
import csv
from itertools import zip_longest
import pandas as pd

curr_dir = Path(__file__).parent.parent
code_dir = curr_dir.parent
app_dir = code_dir.parent
BOM_dir = app_dir / "BOM"
temp_dir = BOM_dir / "template"
bom_path = BOM_dir / "bom_ready.txt"
excel_path = BOM_dir / "bom_ready.xlsx"


def main(main_lines, extension_lines):
    combined_lines = []
    space = "                              ,                                 ,                               ,                           ,"

    i = 0
    for main, ext in zip_longest(main_lines, extension_lines):
        new_line = (main.strip() if main else space) + " , " + (ext.strip() if ext else " ") + "\n"

        i = i + 1
        combined_lines.append(new_line)

    with open(bom_path, "w", encoding="utf-8") as f:
        f.writelines(combined_lines)

    with open(bom_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # First line is the header
    header = [col.strip() for col in lines[0].split(',')]
    rows = []

    for line in lines[1:]:
        # Use csv reader to correctly parse quoted commas
        reader = csv.reader([line], skipinitialspace=True)
        parsed_row = next(reader)
        cleaned_row = [col.strip() for col in parsed_row]
        rows.append(cleaned_row)

    # Convert to DataFrame
    df = pd.DataFrame(rows, columns=header)

    # Save to Excel
    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for column in ws.columns:
        max_length = 0
        col = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(cell.value.strip()))
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
            except:
                pass

            adjustment_width = max_length + 2
            ws.column_dimensions[col].width = adjustment_width

    last_col = ws.max_column
    ws.delete_cols(last_col)

    for row in ws.iter_rows(min_row=2):  # skip header
        cell = row[3]  # colum D = ILOSC
        try:
            value = int(cell.value)
            cell.value = value
        except (ValueError, TypeError):
            pass

    wb.save(excel_path)
