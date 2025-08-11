import pathlib
import pandas as pd
import os
from openpyxl import load_workbook

del_counter = 0


def remove_last_comma(line):
    return line[:-1]


def remove_first_comma(line):
    return line[1:]


def parse_fixed_width_line(line):
    # Extract fixed-width fields
    field1 = line[1:33].strip()
    field2 = line[35:67].strip()
    field3 = line[69:108].strip()
    field4 = line[110:114].strip()
    field5 = line[117:142].strip()
    return field1, field2, field3, field4, field5


def merge_two_lines(curr_result, next_lines_table):  # merge next_line and curr_line
    next_line = next_lines_table[0]
    next_result = parse_fixed_width_line(next_line)

    for i in range(0, 5):
        print(i + 1, next_result[i])
        if next_result[i].strip() != "":
            curr_result[i] = curr_result[i].strip() + " " + next_result[i].strip()

    return curr_result


def format_row_to_fixed_width(row):  # save fileds from table to one line
    widths = [32, 32, 38, 5, 26]

    padded_fields = [
        row[i].strip().ljust(widths[i]) for i in range(min(len(row), len(widths)))
    ]

    return "` ".join(padded_fields)


def mod_long_name(curr_line, next_line, lines, index):
    global del_counter
    line = curr_line
    curr_line = remove_last_comma(curr_line)
    curr_line = remove_first_comma(curr_line)

    curr_line_table = curr_line.split(",")
    next_line_table = next_line.split(",")

    if (
        len(next_line_table) < 4 and len(next_line_table) > 1
    ):  # too long name in next row
        curr_result = list(parse_fixed_width_line(curr_line))
        next_result = list(parse_fixed_width_line(next_line))

        curr_line_table = merge_two_lines(
            curr_result, next_line_table
        )  # merge curr_line and next_line
        curr_line = format_row_to_fixed_width(curr_line_table)

        del lines[index]
        del_counter += 1
        return " ` " + curr_line + " `" + "\n"
    return line


def txt_to_single_excel(txt_path, excel_path):
    with open(txt_path, "r", encoding="utf-8-sig") as f:
        lines = f.readlines()

    # Split each line by commas into a list of cells
    data = []
    for line in lines:
        stripped = line.strip()
        if stripped:  # skip empty lines
            row = [cell.strip() for cell in stripped.split("`")]
            data.append(row)

    # Convert to DataFrame and write to Excel
    df = pd.DataFrame(data)
    df.to_excel(excel_path, index=False, header=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for column in ws.columns:
        max_length = 0
        col = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(cell.value.strip()))
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
            except:
                pass

            adjustment_width = max_length + 2
            ws.column_dimensions[col].width = adjustment_width

    for row in ws.iter_rows(min_row=2):  # skip header
        cell = row[4]  # colum D = ILOSC
        try:
            value = int(cell.value)
            cell.value = value
        except (ValueError, TypeError):
            pass

    wb.save(excel_path)


def main(BOM_path, Excel_path, BOM_ready):
    global del_counter

    lines = []
    with open(BOM_path, "r", encoding="utf-8") as f:
        lines = f.readlines()
        i = 0
        while i < len(lines):
            if not lines[i].__contains__("Assembly"):
                if len(lines[i].strip()) == 0:
                    i += 1
                    continue
                if i < len(lines) - 1:
                    lines[i] = mod_long_name(lines[i], lines[i + 1], lines, i)
            i += 1  # move to next line only if not deleted (otherwise mod_long_name deletes next line)

    with open(BOM_ready, "w", encoding="utf-8") as f:
        f.writelines(lines)

    txt_to_single_excel(BOM_ready, Excel_path)


if __name__ == "__main__":
    main(
        Excel_path=r"C:\Users\JakubFornal\Desktop\CreoMate\Zamówienia CreoMate.xlsx",
        BOM_path=r"C:\Users\JakubFornal\Downloads\im32_00000000-montaz_poduszki-z.bom (1).4",
        BOM_ready=r"C:\Users\JakubFornal\Desktop\CreoMate\readyBOM.txt",
    )
