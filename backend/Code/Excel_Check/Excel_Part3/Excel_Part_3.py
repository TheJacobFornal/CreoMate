from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill


wrong_counter = 0

def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill



def main(Excel_path):
    global wrong_counter
    wrong_counter = 0
    Name_Index = 3
    Type_Index = 5
    Creo_Index = 1
    Uwagi_Index = 9
    wb = load_workbook(Excel_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        temp = ws.cell(row, Uwagi_Index ).value
        if temp is not None:

            temp = temp.strip().lower()
            if any(char.isdigit() for char in temp) :
                continue
            else:
                if temp.__contains__("kat."):                             #usuwa ".kat"
                    temp = temp.replace("kat.", "")

                temp= temp[0].upper() + temp[1:]
                temp = temp.strip()

                temp = temp.capitalize()

                ws.cell(row, Uwagi_Index).value = temp
               
    wb.save(Excel_path)


    return wrong_counter 











