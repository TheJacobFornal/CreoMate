import sys
import io
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import re
from itertools import zip_longest

# Ensure UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")


def clean_illegal_chars(val):
    """Remove illegal Excel characters"""
    if isinstance(val, str):
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", val)
    return val


def check_if_add_sign(line):
    """Determine whether to add a backtick separator"""
    line = line[-5:]
    return not ("`" in line)


def normalize_row(row, n_cols):
    """Make sure each row has exactly n_cols elements"""
    if len(row) < n_cols:
        # Pad missing columns
        row += [""] * (n_cols - len(row))
    elif len(row) > n_cols:
        # Merge extra columns into the last column
        row = row[: n_cols - 1] + [" ".join(row[n_cols - 1 :])]
    return row


def main(main_lines, extension_lines, Excel_path, readyBom_path):
    combined_lines = []
    space = "`"  # minimal placeholder if a line is missing

    okey = True

    add_sign = check_if_add_sign(main_lines[2])
    sign = "`"

    # Combine main and extension lines
    for main, ext in zip_longest(main_lines, extension_lines):
        if not main or not ext:
            okey = False

        main_clean = main.strip().rstrip("`") if main else ""
        ext_clean = ext.strip().rstrip("`") if ext else ""

        sep = sign if add_sign else " "
        new_line = f"{main_clean}{sep}{ext_clean}\n"
        combined_lines.append(new_line)

    # Write combined file
    with open(readyBom_path, "w", encoding="utf-8") as f:
        f.writelines(combined_lines)

    # Read combined lines
    with open(readyBom_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Header processing
    header = [col.strip() for col in lines[0].split("`")]
    n_cols = len(header)
    rows = []

    # Process data rows
    for line in lines[1:]:
        parsed_row = [col.strip() for col in line.split("`")]
        cleaned_row = normalize_row(parsed_row, n_cols)
        rows.append(cleaned_row)

    # Create DataFrame
    df = pd.DataFrame(rows, columns=header)

    # Clean illegal characters
    df = df.applymap(clean_illegal_chars)

    # Save to Excel
    df.to_excel(Excel_path, index=False)

    # Open Excel for formatting
    wb = load_workbook(Excel_path)
    ws = wb.active

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        col = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value).strip()))
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2

    # Convert ILOSC column (assumes 4th column, D)
    for row in ws.iter_rows(min_row=2):
        cell = row[3]
        try:
            cell.value = int(cell.value)
        except (ValueError, TypeError):
            pass

    wb.save(Excel_path)
    return okey