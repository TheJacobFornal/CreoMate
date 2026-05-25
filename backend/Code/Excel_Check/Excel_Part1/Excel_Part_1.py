import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path
from Code.Excel_Check.Excel_Addision import Excel_addition

# ===== Constants =====
TAG_COLORS = {
    "00FFB7",
    "FFFF00",
    "FF0000",
    "F76700",
    "ABA200",
    "D3A6FF",
    "00B0F0",
    "A1887F",
    "B0BEC5",
    "FF3399",
    "42FF48",
    "379392",
    "DDD8B8",
    "6699FF",
}
#VALID_TYPES = {"H", "P", "L", "F", "T", "W", "O", "S", "N", "D"}
#VALID_H_TYPES = {f"H{i}" for i in range(1, 15)}
#VALID_CREO_SUFFIX = VALID_TYPES.union(VALID_H_TYPES)

VALID_HANDOLOWE = {
    "HA",
    "SE",
    "CZ",
    "SN",
    "PL",
    "HM",
    "IO",
    "PW",
    "KK",
    "MA",
    "NO",
    "OS",
    "PS",
    "PR",
    "RO",
    "HS",
    "SW",
    "ST",
    "HU",
    "PN",
}

VALID_PRODUCTION = {
    "D",
    "F",
    "L",
    "P",
    "T",
    "U",
    "W",
}

VALID_OTHER ={
    "Z",
    "S"
}

def is_cell_colored(cell):
    """
    Returns RGB hex string of the cell fill (e.g., 'FF0000'), or None if no color.
    Ignores theme colors.
    """
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None

    color = fill.start_color

    # Safely get RGB string
    if color.type == "rgb" and color.rgb:
        rgb_str = str(color.rgb).upper()
        return True
    else:
        return False


# ===== Globals =====
counter_wrong = 0
max_row = 0
min_row = 0


# ===== Utility Functions =====
def contains_exactly_one_letter(s: str) -> bool:
    """Checks if a string contains exactly one letter."""
    return len(re.findall(r"[A-Za-z]", s)) == 1


def contain_num(text: str) -> bool:
    """Checks if a string contains a numeric character."""
    return any(char.isdigit() for char in text)


def color_row(ws, row_num: int, highlight: bool, color: str = "FFFF00"):
    fill = (
        PatternFill(start_color=color, end_color=color, fill_type="solid")
        if highlight
        else PatternFill(fill_type=None)
    )
    for col in range(1, 11):
        ws.cell(row=row_num, column=col).fill = fill


def check_if_colored(cell) -> bool:
    """Checks if a cell is colored with a known tag color."""  # check if cell is colored with a known tag color - zakupy tag
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return False
    fg_color = fill.start_color.rgb
    if not fg_color:
        return False
    if len(fg_color) == 8:  # ARGB
        fg_color = fg_color[-6:]
    return fg_color.upper() not in TAG_COLORS


# ===== Checks =====
def remove_dash_Type(ws, removeHItems: bool):
    """Removes dashes in Type column and highlights or deletes rows with 'H' types."""
    global counter_wrong
    TypeIndex = 5
    global max_row, min_row

    for row in range(max_row, min_row - 1, -1):
        Type_value = ws.cell(row, TypeIndex).value
        if Type_value:
            Type_value = str(Type_value).strip()  # remove "-" from type col.
            removed_dash = Type_value.replace("-", "").upper()
            ws.cell(row, TypeIndex).value = removed_dash.upper()

            if (isinstance(removed_dash, str)  
                and len(removed_dash) == 2
                and removed_dash[0].isalpha()
                and removed_dash[1].isdigit()):
                 
                if removeHItems:
                    ws.delete_rows(row)
                    continue
                color_row(ws, row, True, "FFFF00")  # yellow
                counter_wrong += 1
                
                print("yellow: ", removed_dash)

                continue
            
            elif (removed_dash in VALID_OTHER):
                continue

            elif (removed_dash not in VALID_HANDOLOWE and removed_dash not in VALID_PRODUCTION and removed_dash not in VALID_OTHER) and not is_cell_colored(ws.cell(row, 1)):
                color_row(ws, row, True, "00FFB7")  # aqua
                print("aqua 1: ", removed_dash)
                counter_wrong += 1
            

        elif not is_cell_colored(ws.cell(row, 1)):
            color_row(ws, row, True, "00FFB7")  # aqua
            print("aqua 2: ", removed_dash)
            
            counter_wrong += 1


def check_type_creo_name(ws, row: int, creo: str):
    """Validates that the Creo name has a proper type suffix."""
    global counter_wrong
    if not creo or "-" not in creo:
        return

    last_part = creo.split("-")[-1].strip()

        
    if (last_part not in VALID_HANDOLOWE and last_part not in VALID_PRODUCTION and last_part not in VALID_OTHER) and not is_cell_colored(ws.cell(row, 1)):  # check creo name suffix like IL40_04020004A-KRAZEK-T (-T)
        counter_wrong += 1
        color_row(ws, row, True, "00FFB7")
        print("creo names: ", last_part)


def check_handlowe(ws, row: int):
    """Highlights issues in handlowe rows. 'H'"""
    global counter_wrong
    Producent_index, Nazwa_index, Numer_index = 9, 3, 2

    Producent_cell = ws.cell(row, Producent_index)
    Nazwa_val = ws.cell(row, Nazwa_index).value
    Numer_val = ws.cell(row, Numer_index).value

    if (
        Nazwa_val
        and str(Nazwa_val).strip().lower()
        == "nazwa katalogowa"  # default values - nie uzupełnione relacje
        or Numer_val
        and str(Numer_val).strip().lower() == "nr. katalogowy"
        or Producent_cell
        and Producent_cell.value == "Producent"
    ) and not is_cell_colored(ws.cell(row, 1)):
        color_row(ws, row, True, "F76700")
        counter_wrong += 1
        return

    if (
        not Producent_cell.value
        or Nazwa_val is None
        or Numer_val is None
        and not is_cell_colored(ws.cell(row, 1))
    ):  # check if there is producent, nazwa, numer
        color_row(ws, row, True, "FF0000")  # red
        counter_wrong += 1


def check_Material_Obrobki_brak(
    ws, row: int
):  # change brak -> brak / none in columns 7, 8
    """Normalizes 'Brak' values in Obróbki columns."""
    Cieplna_index, Powierzchnia_index = 7, 8

    if (
        ws.cell(row, Cieplna_index).value
        and str(ws.cell(row, Cieplna_index).value).strip().lower() == "brak"
    ):
        ws.cell(row, Cieplna_index).value = "Brak / None"

    if (
        ws.cell(row, Powierzchnia_index).value
        and str(ws.cell(row, Powierzchnia_index).value).strip().lower() == "brak"
    ):
        ws.cell(row, Powierzchnia_index).value = "Brak / None"


def check_production(
    ws, row: int
):  # sprawdza obróbki i materiał - col. 6, 7, 8 musi być uzuepłniona
    """Highlights missing material or obróbki in production rows."""
    global counter_wrong
    Material_index, Cieplna_index, Powierzchnia_index = 6, 7, 8

    if not (
        ws.cell(row, Material_index).value
        and ws.cell(row, Cieplna_index).value
        and ws.cell(row, Powierzchnia_index).value
    ) and not is_cell_colored(ws.cell(row, 1)):
        color_row(ws, row, True, "ABA200")
        counter_wrong += 1


# ===== Main functions =====
def main_Tree(Excel_path: Path):
    wb = load_workbook(Excel_path)
    ws = wb.active
    producent_index = 6

    for row in range(ws.max_row, 0, -1):
        creo_name = ws.cell(row, 2).value
        if not creo_name:
            continue

        check_type_creo_name(
            ws, row, creo_name
        )  # IL40_04020004A-KRAZEK-T - typ w nazwie creo

        producent = ws.cell(row, producent_index).value

        if producent:  # modyfikcja producenta - usuwa KAT
            producent = producent.replace("kat.", "").strip().capitalize()
            producent = producent.replace("kat", "").strip().capitalize()
            ws.cell(row, producent_index).value = producent

        typ = creo_name.split("-")[-1]
        if typ == "H" or typ == "N":
            if contain_num(typ):
                ws.delete_rows(row)
            elif not producent:
                color_row(ws, row, True)

    wb.save(Excel_path)


def check_clear_Number(ws, row):
    global counter_wrong
    Numer_val = ws.cell(row, 2).value
    if (
        isinstance(Numer_val, str)
        and "nr kat" in Numer_val.lower()
        and not is_cell_colored(ws.cell(row, 1))
    ):  # "nr kat" w numerze
        color_row(ws, row, True, "6699FF")
        counter_wrong += 1


def main(Excel_path: Path, removeHItems=False, Zakupy=False) -> int:
    global counter_wrong, max_row, min_row
    counter_wrong = 0
    TypeIndex = 5

    wb = load_workbook(Excel_path)
    ws = wb.active
    max_row, min_row = Excel_addition.get_max_min_row(ws, Zakupy)

    remove_dash_Type(ws, removeHItems)  # usuwa dash z Typu kol. 5

    for row in range(min_row, ws.max_row + 1):

        if not Zakupy:
            creo_name = ws.cell(row, 1).value
            check_type_creo_name(
                ws, row, creo_name
            )  # IL40_04020004A-KRAZEK-T - typ w nazwie creo
        # else:
        # if not check_if_colored(ws.cell(row, 1)):
        # highlight_repeated_in_column(ws, 2)

        check_clear_Number(ws, row)  # Check for "nr kat" in Numer column

        Type_value = ws.cell(row, TypeIndex).value
        check_Material_Obrobki_brak(ws, row)

        if Type_value:  # check production and handlowe
            if Type_value in VALID_HANDOLOWE:
                check_handlowe(ws, row)
            elif Type_value in VALID_PRODUCTION:
                check_production(ws, row)

    wb.save(Excel_path)
    return counter_wrong


if __name__ == "__main__":
    main(
        Path(r"D:\Creo_Ustawienia\Programiki\CreoMate\Zamówienia CreoMate.xlsx"),
        True,
        Zakupy=True,
    )
    
    
