from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill
from Code.Excel_Check.Excel_Addision import Excel_addition
import os

TAG_COLORS = {
    "00FFB7",  # Mint Aqua
    "FFFF00",  # Bright Yellow
    "FF0000",  # Red
    "F76700",  # Orange
    "ABA200",  # Olive
    "D3A6FF",  # Lavender Violet
    "00B0F0",  # Cyan Blue
    "A1887F",  # Light Brown / Taupe
    "A1A1A1",  # Cool Gray
    "FF0095",  # Hot Pink
    "42FF48",  # Neon Green
    "DDD8B8",  # sandy,
    "379392",  # Dark syjan
    "DDD8B8", # sandy,
    "6699FF",   # Light Blue
}
VALID_TYPES = {'H', 'P', 'L', 'F', 'T', 'W', 'O', 'S', 'N', 'D'}

def is_cell_colored(cell):
    """
    Returns RGB hex string of the cell fill (e.g., 'FF0000'), or None if no color.
    Ignores theme colors.
    """
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None

    color = fill.start_color

    # Safely get RGB string
    if color.type == "rgb" and color.rgb:
        rgb_str = str(color.rgb).upper()
        return True
    else:
        return False




min_row = 0
max_row = 0
wrong_counter = 0
def color_row(ws, row_num, type, color = "FFFF00"):
    if type:
        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    else:
        fill = PatternFill(fill_type= None)
    for col in range(1, 10):
        ws.cell(row=row_num, column=col).fill = fill

def check_profile(ws, row):                                                 # compare number H and length of profile kol. A wit kol B - grey
    global wrong_counter
    Number_Index = 2
    Creo_Index = 1

    number_val = ws.cell(row, Number_Index).value
    creo_val = ws.cell(row, Creo_Index).value

    number_creo = creo_val.split("-")[0]

    if number_creo != number_val:
        wrong_counter += 1
        color_row(ws, row, True, "A1A1A1")
    else:
        color_row(ws, row, False)

def modify_left_duplicate(ws, row_left, base_name_left, removeMirror):
    global wrong_counter
    global max_row
    global min_row
    
    
    left_quantity = ws.cell(row_left, 4).value
    base_name = base_name_left[:-1]                              # base part of the name left without "L"
    right_find = False
    
    for row in range(min_row, max_row + 1):  # find right component
        creo_right_value = ws.cell(row, 1).value
        print(creo_right_value, base_name, base_name_left, flush=True)
        if (
            creo_right_value
            and base_name in creo_right_value
            and base_name_left not in creo_right_value
        ):  # found right component, not the same as left
           
            right_quantity = ws.cell(row, 4).value

            if "+" not in str(right_quantity):  # not yet merged
                quantity = str(right_quantity).strip() + "+" + str(left_quantity).strip() + "L"
                ws.cell(row, 4).value = quantity  # write new combined quantity

            if removeMirror:
                ws.delete_rows(row_left) 
            else:
                wrong_counter += 1
                color_row(ws, row_left, True, "FF0095")  # mark left as processed - PINK    

            right_find = True
            break  # stop after first match


    if not right_find:                                                                                                       # there isn't right component
        left_quantity = ws.cell(row_left, 4).value

        if not str(left_quantity).__contains__("+"): # no duplicates
            quantity = "0+ " + str(left_quantity).strip() + "L"
            ws.cell(row_left, 4).value = quantity

            Note = "Wykonać w lustrze!"
            ws.cell(row_left, 9).value = Note

        color_row(ws, row_left, True, "42FF48")

def check_if_mirror(creo_name):
    x = creo_name[-1]
    if x == "L":
        print("Left mirror found in name:", creo_name, flush=True)
        return True
    return False
 
def highlight_repeated_in_column(ws, col: int):
    """Highlights rows where column values are repeated."""
    value_count = {}
    global wrong_counter

    for row in range(2, max_row + 1):
        print("row:", row, flush=True)
        if is_cell_colored(ws.cell(row, 1)):
            continue
        typ = ws.cell(row, 5).value
        creo_name = ws.cell(row, 1).value

        if creo_name:
            base_name = creo_name.split('-')[0] 
        
        if typ in VALID_TYPES and not base_name.endswith('L'):
            val = ws.cell(row, col).value
            value_count.setdefault(val, []).append(row)
            
    for value, rows in value_count.items():
        print(f"Value: {value} → Rows: {rows}")


    for rows in value_count.values():
        if len(rows) > 1:
            for row in rows:
                color_row(ws, row, True, "DDD8B8")
                wrong_counter += 1
                print("counter ++ repetion", ws.cell(row, 1).value, flush=True)   


def main(Excel_path, removeMirror, Zakupy=False):
    global wrong_counter
    global max_row
    global min_row
    
    wrong_counter = 0
    Name_Index = 3
    Type_Index = 5
    Creo_Index = 1
    wb = load_workbook(Excel_path)
    ws = wb.active
    
    max_row, min_row = Excel_addition.get_max_min_row(ws, Zakupy)
    for row in range(max_row, min_row - 1, -1):
        Name_value = ws.cell(row, 3).value
        Type_value = ws.cell(row, 5).value
        creo_val = ws.cell(row, 1).value

       
        if Name_value is not None:
            Upper_name = str(Name_value).upper()
            ws.cell(row, 3).value = Upper_name                                            # upper letter in kol. Name
            if Upper_name.__contains__("PROFIL_") and Type_value.__contains__("H") and not is_cell_colored(ws.cell(row, 1)):                # Profile with dimensions Type: H
                if Zakupy:
                    continue
                else:
                    check_profile(ws, row)
            else:
                Numer_val = ws.cell(row, 2).value                                                  # Number contain "_" - brown
                if Numer_val is not None and  str(Numer_val).__contains__("_") and not is_cell_colored(ws.cell(row, 1)):
                    print("Number with underscore found in row:", row, flush=True)
                    wrong_counter += 1
                    color_row(ws, row, True, "D3A6FF")

                if not Zakupy:
                    base_name = creo_val.split("-")[0]                      # IL31_02130302 or IL31_02130302L
                    
                    if check_if_mirror(base_name) and Type_value in {'P', 'L', 'F', 'T', 'W', 'O', 'S', 'D',}:
                        print("Left mirror found in row:", base_name, flush=True)
                        modify_left_duplicate(ws, row, base_name, removeMirror)

       
       
    highlight_repeated_in_column(ws, 2)                                # powtórzenia w kolumnie 2  
    
    print("Counter wrong:", wrong_counter, flush=True)
    
    wb.save(Excel_path)

    return wrong_counter


if __name__ == "__main__":
    mode = input("1 = remove mirror, 2 = do not remove mirror: ")
    if mode == "1":
        removeMirror = True
    else:
        removeMirror = False
        
    main(Path(r"C:\Users\JakubFornal\Desktop\CreoMate\BOM CreoMate.xlsx"), removeMirror, Zakupy=False)
    print("finished")
    os.startfile(Path(r"C:\Users\JakubFornal\Desktop\CreoMate\BOM CreoMate.xlsx"))